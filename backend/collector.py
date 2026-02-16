#!/usr/bin/env python3
"""
BRVM Collector v2 — Calibré sur le Bulletin Officiel de la Cote réel
Structure basée sur boc_20260211_2.pdf (N°29, mercredi 11 février 2026)

Colonnes réelles du tableau "MARCHE DES ACTIONS" :
  Code Sect. | Symbole | Titre | Cours Précédent | Cours Ouv. | Cours Clôt. |
  Variation jour | Volume | Valeur | Cours Référence | Variation annuelle |
  Dernier dividende (Montant net + Date) | Rdt. Net | PER

Compartiments : PRESTIGE (12 titres) + PRINCIPAL (35 titres)
Indices : BRVM COMPOSITE, BRVM 30, BRVM PRESTIGE
"""

import re
import sqlite3
import requests
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, date, timedelta
from pathlib import Path
import logging
import urllib3
import argparse
import json

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── CONFIGURATION ─────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("brvm_collector.log"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "brvm_data.db"
PDF_DIR = BASE_DIR / "bulletins"
EXCEL_DIR = BASE_DIR / "rapports"
PDF_DIR.mkdir(exist_ok=True)
EXCEL_DIR.mkdir(exist_ok=True)

# URL patterns du site BRVM (format observé : boc_YYYYMMDD_2.pdf)
URL_PATTERNS = [
    "https://www.brvm.org/sites/default/files/boc_{date}_2.pdf",
    "https://www.brvm.org/sites/default/files/boc_{date}_1.pdf",
]

# Mapping secteurs (code 2 lettres → libellé complet)
SECTEURS = {
    "TEL": "BRVM-TELECOMMUNICATIONS",
    "FIN": "BRVM-SERVICES FINANCIERS",
    "CD":  "BRVM-CONSOMMATION DISCRETIONNAIRE",
    "CB":  "BRVM-CONSOMMATION DE BASE",
    "IND": "BRVM-INDUSTRIELS",
    "ENE": "BRVM-ENERGIE",
    "SPU": "BRVM-SERVICES PUBLICS",
}

# Tous les symboles BRVM connus (compartiment PRESTIGE + PRINCIPAL)
# Source : boc_20260211_2.pdf
SYMBOLES_BRVM = {
    # COMPARTIMENT PRESTIGE (12 titres)
    "NTLC", "PALC", "SPHC", "SMBC", "TTLC", "TTLS",
    "ECOC", "SGBC", "SIBC", "ONTBF", "ORAC", "SNTS",
    # COMPARTIMENT PRINCIPAL (35 titres)
    "SCRC", "SICC", "SLBC", "SOGC", "STBC", "UNLC",
    "ABJC", "BNBC", "CFAC", "LNBB", "NEIC", "PRSC", "UNXC",
    "SHEC", "BICB", "BICC", "BOAB", "BOABF", "BOAC", "BOAM",
    "BOAN", "BOAS", "CBIBF", "ETIT", "NSBC", "ORGT", "SAFC",
    "CABC", "FTSC", "SDSC", "SEMC", "SIVC", "STAC",
    "CIEC", "SDCC",
}


# ── BASE DE DONNÉES ───────────────────────────────────────────────────────────

def init_db():
    """Crée les tables si elles n'existent pas."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # Table séances
    c.execute("""
        CREATE TABLE IF NOT EXISTS seances (
            date TEXT PRIMARY KEY,
            seance_num INTEGER,
            jour_semaine TEXT,
            composite REAL,
            var_composite REAL,
            var_composite_annuelle REAL,
            brvm30 REAL,
            var_brvm30 REAL,
            var_brvm30_annuelle REAL,
            prestige REAL,
            var_prestige REAL,
            var_prestige_annuelle REAL,
            capitalisation INTEGER,
            volume_total INTEGER,
            valeur_totale INTEGER,
            nb_titres INTEGER,
            nb_hausse INTEGER,
            nb_baisse INTEGER,
            nb_inchange INTEGER,
            nb_societes INTEGER
        )
    """)
    
    # Table cours des actions (colonnes réelles du bulletin)
    c.execute("""
        CREATE TABLE IF NOT EXISTS cours (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            compartiment TEXT,       -- PRESTIGE ou PRINCIPAL
            secteur_code TEXT,       -- CB, FIN, IND, etc.
            secteur_libelle TEXT,    -- BRVM-SERVICES FINANCIERS, etc.
            symbole TEXT NOT NULL,
            titre TEXT,
            cours_precedent REAL,
            cours_ouverture REAL,
            cours_cloture REAL,
            variation_jour REAL,
            volume INTEGER,
            valeur_seance INTEGER,
            cours_reference REAL,
            variation_annuelle REAL,
            dividende_montant REAL,
            dividende_date TEXT,
            rendement_net REAL,
            per REAL,
            UNIQUE(date, symbole)
        )
    """)
    
    # Table indices sectoriels
    c.execute("""
        CREATE TABLE IF NOT EXISTS indices_sectoriels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            secteur_code TEXT,
            nb_societes INTEGER,
            valeur REAL,
            var_jour REAL,
            var_annuelle REAL,
            volume INTEGER,
            valeur_echangee INTEGER,
            per_moyen REAL,
            UNIQUE(date, secteur_code)
        )
    """)
    
    # Table conseils d'investissement
    c.execute("""
        CREATE TABLE IF NOT EXISTS conseils (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date_conseil TEXT,
            symbole TEXT,
            titre TEXT,
            type TEXT CHECK(type IN ('ACHAT', 'VENTE', 'NEUTRE')),
            prix_entree REAL,
            prix_cible REAL,
            stop_loss REAL,
            commentaire TEXT,
            actif INTEGER DEFAULT 1,
            date_cloture TEXT,
            resultat_pct REAL
        )
    """)
    
    conn.commit()
    conn.close()
    log.info("Base de données initialisée")


# ── TÉLÉCHARGEMENT PDF ────────────────────────────────────────────────────────

def download_bulletin(target_date: date, force: bool = False) -> Path | None:
    """Télécharge le bulletin PDF pour une date donnée."""
    date_str = target_date.strftime("%Y%m%d")
    pdf_path = PDF_DIR / f"boc_{date_str}.pdf"
    
    if pdf_path.exists() and not force:
        log.info(f"Bulletin déjà téléchargé : {pdf_path.name}")
        return pdf_path
    
    session = requests.Session()
    session.verify = False
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    })
    
    for pattern in URL_PATTERNS:
        url = pattern.format(date=date_str)
        try:
            log.info(f"Tentative : {url}")
            resp = session.get(url, timeout=30)
            if resp.status_code == 200 and len(resp.content) > 5000:
                pdf_path.write_bytes(resp.content)
                log.info(f"✓ Bulletin téléchargé : {pdf_path.name} ({len(resp.content):,} octets)")
                return pdf_path
            else:
                log.warning(f"Réponse invalide : {resp.status_code}")
        except Exception as e:
            log.warning(f"Erreur : {e}")
    
    log.error(f"Bulletin introuvable pour {target_date}")
    return None


# ── EXTRACTION PDF ────────────────────────────────────────────────────────────

def parse_float(s: str) -> float | None:
    """Convertit une chaîne en float (gère virgules, espaces, %)."""
    if not s or s.strip() in ("", "-", "NC", "ND", "SP"):
        return None
    s = str(s).strip()
    s = s.replace("\xa0", "").replace(" ", "").replace(",", ".")
    s = s.replace("%", "").replace("+", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_int(s: str) -> int | None:
    """Convertit une chaîne en entier."""
    v = parse_float(s)
    return int(v) if v is not None else None


def extract_page1_data(pdf) -> dict:
    """
    Extrait les données de la page 1 :
    - Indices BRVM COMPOSITE, BRVM 30, BRVM PRESTIGE
    - Statistiques du marché
    - Date et numéro de séance
    """
    result = {}
    page = pdf.pages[0]
    text = page.extract_text() or ""
    
    # ── Date et numéro de séance
    date_match = re.search(r'(\w+)\s+(\d{1,2})\s+([\wéû]+)\s+(\d{4})', text)
    num_match = re.search(r'N°\s*(\d+)', text)
    
    if date_match:
        result["date_texte"] = f"{date_match.group(1)} {date_match.group(2)} {date_match.group(3)} {date_match.group(4)}"
    if num_match:
        result["seance_num"] = int(num_match.group(1))
    
    # ── BRVM COMPOSITE
    comp_match = re.search(r'BRVM\s+COMPOSITE\s+([\d\s,\.]+)', text)
    if comp_match:
        result["composite"] = parse_float(comp_match.group(1))
    
    # Variations COMPOSITE
    var_matches = re.findall(r'Variation\s+Jour\s+([-+]?[\d,\.]+)\s*%', text)
    vann_matches = re.findall(r'Variation\s+annuelle\s+([-+]?[\d,\.]+)\s*%', text)
    
    if len(var_matches) >= 1:
        result["var_composite"] = parse_float(var_matches[0])
    if len(var_matches) >= 2:
        result["var_brvm30"] = parse_float(var_matches[1])
    if len(var_matches) >= 3:
        result["var_prestige"] = parse_float(var_matches[2])
    
    if len(vann_matches) >= 1:
        result["var_composite_annuelle"] = parse_float(vann_matches[0])
    if len(vann_matches) >= 2:
        result["var_brvm30_annuelle"] = parse_float(vann_matches[1])
    if len(vann_matches) >= 3:
        result["var_prestige_annuelle"] = parse_float(vann_matches[2])
    
    # ── BRVM 30
    brvm30_match = re.search(r'BRVM\s+30\s+([\d,\.]+)', text)
    if brvm30_match:
        result["brvm30"] = parse_float(brvm30_match.group(1))
    
    # ── BRVM PRESTIGE
    pres_match = re.search(r'BRVM\s+PRESTIGE\s+([\d,\.]+)', text)
    if pres_match:
        result["prestige"] = parse_float(pres_match.group(1))
    
    # ── Statistiques marché (Actions)
    # Capitalisation
    cap_match = re.search(r'Capitalisation\s+bours[iè]+re\s*\(FCFA\)[^\d]*([\d\s]+)', text)
    if cap_match:
        result["capitalisation"] = parse_int(cap_match.group(1).replace(" ", ""))
    
    # Volume
    vol_match = re.search(r'Volume\s+échangé\s*\(Actions[^\)]*\)\s*([\d\s]+)', text)
    if vol_match:
        result["volume_total"] = parse_int(vol_match.group(1).replace(" ", ""))
    
    # Valeur transigée
    val_match = re.search(r'Valeur\s+trans[iig]+ée\s*\(FCFA\)\s*\(Actions[^\)]*\)\s*([\d\s]+)', text)
    if val_match:
        result["valeur_totale"] = parse_int(val_match.group(1).replace(" ", ""))
    
    # Nombre de titres
    nb_match = re.search(r'Nombre\s+de\s+titres\s+transigés\s+(\d+)', text)
    if nb_match:
        result["nb_titres"] = int(nb_match.group(1))
    
    hausse_match = re.search(r'Nombre\s+de\s+titres\s+en\s+hausse\s+(\d+)', text)
    if hausse_match:
        result["nb_hausse"] = int(hausse_match.group(1))
    
    baisse_match = re.search(r'Nombre\s+de\s+titres\s+en\s+baisse\s+(\d+)', text)
    if baisse_match:
        result["nb_baisse"] = int(baisse_match.group(1))
    
    inchange_match = re.search(r'Nombre\s+de\s+titres\s+inchang[eé]s\s+(\d+)', text)
    if inchange_match:
        result["nb_inchange"] = int(inchange_match.group(1))
    
    log.info(f"Page 1 extraite : Composite={result.get('composite')}, BRVM30={result.get('brvm30')}, "
             f"Prestige={result.get('prestige')}, Séance N°{result.get('seance_num')}")
    return result


def extract_actions(pdf) -> list[dict]:
    """
    Extrait les données des actions depuis les pages 3-4 du bulletin.
    
    Structure réelle du tableau :
    Code Sect. | Symbole | Titre | Cours Précédent | Cours Ouv. | Cours Clôt. |
    Variation jour | Volume | Valeur | Cours Référence | Variation annuelle | ...
    """
    actions = []
    compartiment_actuel = None
    
    for page_num in [2, 3]:  # Pages 3 et 4 (index 2 et 3)
        if page_num >= len(pdf.pages):
            continue
        
        page = pdf.pages[page_num]
        text = page.extract_text() or ""
        
        # Détecter le changement de compartiment
        if "COMPARTIMENT PRESTIGE" in text:
            compartiment_actuel = "PRESTIGE"
        if "COMPARTIMENT PRINCIPAL" in text:
            # Les deux peuvent apparaître sur la même page
            pass
        
        # Extraire les tables
        tables = page.extract_tables()
        
        for table in tables:
            if not table or len(table) < 2:
                continue
            
            for row in table:
                if not row or len(row) < 6:
                    continue
                
                # Chercher si la ligne contient un symbole BRVM connu
                symbole_trouve = None
                secteur_trouve = None
                
                for i, cell in enumerate(row):
                    cell_str = str(cell or "").strip().upper()
                    
                    # Symbole = 3-5 lettres majuscules dans SYMBOLES_BRVM
                    if cell_str in SYMBOLES_BRVM:
                        symbole_trouve = cell_str
                        # Le secteur est souvent dans la cellule précédente
                        if i > 0:
                            prev = str(row[i-1] or "").strip().upper()
                            if prev in SECTEURS:
                                secteur_trouve = prev
                            elif i > 1:
                                prev2 = str(row[i-2] or "").strip().upper()
                                if prev2 in SECTEURS:
                                    secteur_trouve = prev2
                        break
                
                if not symbole_trouve:
                    continue
                
                # Mise à jour compartiment selon contexte
                row_text = " ".join(str(c or "") for c in row)
                
                # Extraire les valeurs numériques
                # L'ordre des colonnes selon le bulletin :
                # [sect_code] [symbole] [titre] [cours_prec] [cours_ouv] [cours_clot] 
                # [var_jour%] [volume] [valeur] [cours_ref] [var_annuelle%] [div_montant] [div_date] [rdt%] [per]
                
                nums = []
                titre = ""
                for cell in row:
                    cell_str = str(cell or "").strip()
                    if not cell_str or cell_str in SYMBOLES_BRVM or cell_str.upper() in SECTEURS:
                        continue
                    v = parse_float(cell_str)
                    if v is not None:
                        nums.append(v)
                    elif len(cell_str) > 3 and not cell_str.replace(" ", "").isnumeric():
                        titre = cell_str
                
                # On a besoin d'au moins cours_prec, cours_clot, variation
                if len(nums) < 3:
                    continue
                
                action = {
                    "compartiment": compartiment_actuel,
                    "secteur_code": secteur_trouve,
                    "secteur_libelle": SECTEURS.get(secteur_trouve, ""),
                    "symbole": symbole_trouve,
                    "titre": titre,
                    "cours_precedent": nums[0] if len(nums) > 0 else None,
                    "cours_ouverture": nums[1] if len(nums) > 1 else None,
                    "cours_cloture": nums[2] if len(nums) > 2 else None,
                    "variation_jour": nums[3] if len(nums) > 3 else None,
                    "volume": int(nums[4]) if len(nums) > 4 else None,
                    "valeur_seance": int(nums[5]) if len(nums) > 5 else None,
                    "cours_reference": nums[6] if len(nums) > 6 else None,
                    "variation_annuelle": nums[7] if len(nums) > 7 else None,
                    "dividende_montant": nums[8] if len(nums) > 8 else None,
                    "rendement_net": nums[9] if len(nums) > 9 else None,
                    "per": nums[10] if len(nums) > 10 else None,
                }
                
                # Correction : variation_jour est en % avec signe
                # Si le cours clôt > cours prec : positif, sinon négatif
                if action["cours_cloture"] and action["cours_precedent"] and action["variation_jour"] is None:
                    action["variation_jour"] = round(
                        (action["cours_cloture"] - action["cours_precedent"]) / action["cours_precedent"] * 100, 2
                    )
                
                actions.append(action)
    
    # Méthode alternative : extraction par regex sur le texte brut
    # (plus robuste pour les PDFs avec mise en page complexe)
    if len(actions) < 5:
        actions = extract_actions_regex(pdf)
    
    log.info(f"Actions extraites : {len(actions)} titres")
    return actions


def extract_actions_regex(pdf) -> list[dict]:
    """
    Méthode alternative d'extraction par regex sur le texte brut.
    Utilisée si l'extraction par tableau échoue.
    """
    actions = []
    compartiment_actuel = "PRESTIGE"
    
    for page_num in [2, 3]:
        if page_num >= len(pdf.pages):
            continue
        
        page = pdf.pages[page_num]
        text = page.extract_text() or ""
        lines = text.split("\n")
        
        for line in lines:
            # Détecter changement de compartiment
            if "COMPARTIMENT PRESTIGE" in line.upper():
                compartiment_actuel = "PRESTIGE"
            elif "COMPARTIMENT PRINCIPAL" in line.upper():
                compartiment_actuel = "PRINCIPAL"
            
            # Chercher une ligne avec un symbole BRVM
            words = line.split()
            if not words:
                continue
            
            symbole = None
            secteur = None
            
            for i, word in enumerate(words):
                w = word.strip().upper().rstrip("*")
                if w in SYMBOLES_BRVM:
                    symbole = w
                    # Chercher secteur dans les mots précédents
                    for j in range(max(0, i-2), i):
                        if words[j].strip().upper() in SECTEURS:
                            secteur = words[j].strip().upper()
                            break
                    break
            
            if not symbole:
                continue
            
            # Extraire les nombres de la ligne
            # Pattern : nombres entiers ou décimaux, positifs ou négatifs, avec ou sans %
            nombre_pattern = r'[-+]?[\d\s]+[,\.]?[\d]*'
            
            # Extraire tous les chiffres de la ligne
            nums_raw = re.findall(r'[-+]?[\d]+(?:[\s][\d]{3})*(?:[,\.][\d]+)?(?:\s*%)?', line)
            nums = []
            for n in nums_raw:
                v = parse_float(n.replace(" ", ""))
                if v is not None:
                    nums.append(v)
            
            if len(nums) < 3:
                continue
            
            # Extraire le titre (texte entre symbole et premiers chiffres)
            sym_pos = line.upper().find(symbole)
            if sym_pos >= 0:
                after_sym = line[sym_pos + len(symbole):]
                titre_match = re.match(r'\s+([A-ZÉÈÊÀÂÎÏÔÙÛÜ\'\s\(\)]+?)(?=\s+[\d])', after_sym)
                titre = titre_match.group(1).strip() if titre_match else ""
            else:
                titre = ""
            
            action = {
                "compartiment": compartiment_actuel,
                "secteur_code": secteur,
                "secteur_libelle": SECTEURS.get(secteur, ""),
                "symbole": symbole,
                "titre": titre,
                "cours_precedent": nums[0] if len(nums) > 0 else None,
                "cours_ouverture": nums[1] if len(nums) > 1 else None,
                "cours_cloture": nums[2] if len(nums) > 2 else None,
                "variation_jour": nums[3] if len(nums) > 3 else None,
                "volume": int(nums[4]) if len(nums) > 4 and nums[4] < 1e9 else None,
                "valeur_seance": int(nums[5]) if len(nums) > 5 else None,
                "cours_reference": nums[6] if len(nums) > 6 else None,
                "variation_annuelle": nums[7] if len(nums) > 7 else None,
                "dividende_montant": None,
                "rendement_net": None,
                "per": None,
            }
            
            if symbole not in [a["symbole"] for a in actions]:
                actions.append(action)
    
    return actions


def extract_date_from_pdf(pdf) -> date | None:
    """Extrait la date de séance depuis le bulletin."""
    mois = {
        "janvier": 1, "février": 2, "mars": 3, "avril": 4,
        "mai": 5, "juin": 6, "juillet": 7, "août": 8,
        "septembre": 9, "octobre": 10, "novembre": 11, "décembre": 12
    }
    
    page = pdf.pages[0]
    text = page.extract_text() or ""
    
    # Pattern : "mercredi 11 février 2026"
    match = re.search(
        r'(?:lundi|mardi|mercredi|jeudi|vendredi)\s+(\d{1,2})\s+'
        r'(janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre)\s+'
        r'(\d{4})',
        text, re.IGNORECASE
    )
    
    if match:
        jour = int(match.group(1))
        mois_num = mois.get(match.group(2).lower(), 0)
        annee = int(match.group(3))
        if mois_num:
            return date(annee, mois_num, jour)
    
    return None


# ── SAUVEGARDE EN BASE ─────────────────────────────────────────────────────────

def save_seance(conn, date_str: str, page1: dict):
    """Sauvegarde les données de séance."""
    c = conn.cursor()
    try:
        c.execute("""
            INSERT OR REPLACE INTO seances 
            (date, seance_num, composite, var_composite, var_composite_annuelle,
             brvm30, var_brvm30, var_brvm30_annuelle,
             prestige, var_prestige, var_prestige_annuelle,
             capitalisation, volume_total, valeur_totale,
             nb_titres, nb_hausse, nb_baisse, nb_inchange)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            date_str,
            page1.get("seance_num"),
            page1.get("composite"),
            page1.get("var_composite"),
            page1.get("var_composite_annuelle"),
            page1.get("brvm30"),
            page1.get("var_brvm30"),
            page1.get("var_brvm30_annuelle"),
            page1.get("prestige"),
            page1.get("var_prestige"),
            page1.get("var_prestige_annuelle"),
            page1.get("capitalisation"),
            page1.get("volume_total"),
            page1.get("valeur_totale"),
            page1.get("nb_titres"),
            page1.get("nb_hausse"),
            page1.get("nb_baisse"),
            page1.get("nb_inchange"),
        ))
        conn.commit()
        log.info(f"Séance {date_str} sauvegardée (N°{page1.get('seance_num')})")
    except Exception as e:
        log.error(f"Erreur sauvegarde séance : {e}")


def save_actions(conn, date_str: str, actions: list[dict]):
    """Sauvegarde les cours des actions."""
    c = conn.cursor()
    inserted = 0
    
    for a in actions:
        try:
            c.execute("""
                INSERT OR IGNORE INTO cours
                (date, compartiment, secteur_code, secteur_libelle, symbole, titre,
                 cours_precedent, cours_ouverture, cours_cloture, variation_jour,
                 volume, valeur_seance, cours_reference, variation_annuelle,
                 dividende_montant, dividende_date, rendement_net, per)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                date_str,
                a.get("compartiment"),
                a.get("secteur_code"),
                a.get("secteur_libelle"),
                a["symbole"],
                a.get("titre"),
                a.get("cours_precedent"),
                a.get("cours_ouverture"),
                a.get("cours_cloture"),
                a.get("variation_jour"),
                a.get("volume"),
                a.get("valeur_seance"),
                a.get("cours_reference"),
                a.get("variation_annuelle"),
                a.get("dividende_montant"),
                a.get("dividende_date"),
                a.get("rendement_net"),
                a.get("per"),
            ))
            inserted += 1
        except Exception as e:
            log.warning(f"Erreur insertion {a.get('symbole')} : {e}")
    
    conn.commit()
    log.info(f"{inserted}/{len(actions)} actions sauvegardées pour {date_str}")
    return inserted


# ── GÉNÉRATION EXCEL ───────────────────────────────────────────────────────────

def style_header(ws, row, cols, bg_color="1B3A6B", font_color="FFFFFF"):
    """Applique le style d'en-tête BRVM (bleu marine + blanc)."""
    fill = PatternFill("solid", fgColor=bg_color)
    font = Font(color=font_color, bold=True, size=10)
    for col_idx, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def generate_excel(target_date: date) -> Path:
    """Génère un rapport Excel complet pour une date donnée."""
    date_str = target_date.strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_PATH)
    
    wb = Workbook()
    
    # ── Feuille 1 : Marché du jour ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = f"Marché {date_str}"
    
    # En-tête
    ws1.merge_cells("A1:Q1")
    ws1["A1"] = f"BRVM — Bulletin Officiel de la Cote — {date_str}"
    ws1["A1"].font = Font(bold=True, size=14, color="1B3A6B")
    ws1["A1"].alignment = Alignment(horizontal="center")
    
    # Indices
    ws1["A2"] = "BRVM COMPOSITE"
    ws1["A3"] = "BRVM 30"
    ws1["A4"] = "BRVM PRESTIGE"
    
    seance = conn.execute("SELECT * FROM seances WHERE date = ?", (date_str,)).fetchone()
    if seance:
        ws1["B2"] = seance[3]  # composite
        ws1["C2"] = f"{seance[4]:+.2f}%" if seance[4] else ""
        ws1["B3"] = seance[6]  # brvm30
        ws1["C3"] = f"{seance[7]:+.2f}%" if seance[7] else ""
        ws1["B4"] = seance[9]  # prestige
        ws1["C4"] = f"{seance[10]:+.2f}%" if seance[10] else ""
    
    # Tableau des actions
    headers = ["Comp.", "Sect.", "Symbole", "Titre", "Cours Préc.",
               "Cours Ouv.", "Cours Clôt.", "Var. Jour %", "Volume",
               "Valeur (FCFA)", "Cours Réf.", "Var. Annuelle %", "Dividende", "Rdt %", "PER"]
    style_header(ws1, 6, headers)
    
    actions = conn.execute("""
        SELECT compartiment, secteur_code, symbole, titre,
               cours_precedent, cours_ouverture, cours_cloture, variation_jour,
               volume, valeur_seance, cours_reference, variation_annuelle,
               dividende_montant, rendement_net, per
        FROM cours WHERE date = ?
        ORDER BY compartiment, secteur_code, symbole
    """, (date_str,)).fetchall()
    
    # Couleurs alternées
    fill_pair = [
        PatternFill("solid", fgColor="EBF3FB"),
        PatternFill("solid", fgColor="FFFFFF"),
    ]
    
    for i, row in enumerate(actions, 7):
        fill = fill_pair[i % 2]
        for j, val in enumerate(row, 1):
            cell = ws1.cell(row=i, column=j, value=val)
            cell.fill = fill
            # Couleurs variation
            if j == 8 and val is not None:  # variation_jour
                if val > 0:
                    cell.font = Font(color="006400", bold=True)
                elif val < 0:
                    cell.font = Font(color="CC0000", bold=True)
    
    # Largeurs colonnes
    for col, width in enumerate([8, 5, 8, 30, 12, 12, 12, 10, 12, 16, 12, 12, 12, 8, 8], 1):
        ws1.column_dimensions[chr(64 + col)].width = width
    
    # ── Feuille 2 : Top Hausses / Baisses ────────────────────────────────────
    ws2 = wb.create_sheet("Pépite & Flop")
    ws2["A1"] = f"TOP HAUSSES & BAISSES — {date_str}"
    ws2["A1"].font = Font(bold=True, size=12, color="1B3A6B")
    
    style_header(ws2, 3, ["Symbole", "Titre", "Cours Clôt.", "Variation Jour", "Volume", "Valeur"])
    ws2["A2"] = "🏆 TOP 5 HAUSSES"
    ws2["A2"].font = Font(bold=True, color="006400")
    
    hausses = conn.execute("""
        SELECT symbole, titre, cours_cloture, variation_jour, volume, valeur_seance
        FROM cours WHERE date = ? AND variation_jour IS NOT NULL
        ORDER BY variation_jour DESC LIMIT 5
    """, (date_str,)).fetchall()
    
    for i, row in enumerate(hausses, 4):
        for j, val in enumerate(row, 1):
            ws2.cell(row=i, column=j, value=val)
    
    ws2["A10"] = "📉 TOP 5 BAISSES"
    ws2["A10"].font = Font(bold=True, color="CC0000")
    style_header(ws2, 11, ["Symbole", "Titre", "Cours Clôt.", "Variation Jour", "Volume", "Valeur"])
    
    baisses = conn.execute("""
        SELECT symbole, titre, cours_cloture, variation_jour, volume, valeur_seance
        FROM cours WHERE date = ? AND variation_jour IS NOT NULL
        ORDER BY variation_jour ASC LIMIT 5
    """, (date_str,)).fetchall()
    
    for i, row in enumerate(baisses, 12):
        for j, val in enumerate(row, 1):
            ws2.cell(row=i, column=j, value=val)
    
    conn.close()
    
    # Sauvegarde
    excel_path = EXCEL_DIR / f"brvm_{date_str}.xlsx"
    wb.save(excel_path)
    log.info(f"Rapport Excel généré : {excel_path}")
    return excel_path


# ── PIPELINE PRINCIPAL ─────────────────────────────────────────────────────────

def process_bulletin(pdf_path: Path, target_date: date) -> dict:
    """Traite un bulletin PDF et retourne un résumé."""
    date_str = target_date.strftime("%Y-%m-%d")
    
    with pdfplumber.open(pdf_path) as pdf:
        log.info(f"PDF ouvert : {pdf_path.name} ({len(pdf.pages)} pages)")
        
        # Extraction
        page1 = extract_page1_data(pdf)
        
        # Essayer d'extraire la date depuis le PDF si pas fournie
        pdf_date = extract_date_from_pdf(pdf)
        if pdf_date and pdf_date != target_date:
            log.warning(f"Date PDF ({pdf_date}) ≠ date demandée ({target_date}), on utilise la date PDF")
            date_str = pdf_date.strftime("%Y-%m-%d")
            target_date = pdf_date
        
        actions = extract_actions(pdf)
    
    # Sauvegarde
    conn = sqlite3.connect(DB_PATH)
    save_seance(conn, date_str, page1)
    nb_actions = save_actions(conn, date_str, actions)
    conn.close()
    
    return {
        "date": date_str,
        "seance_num": page1.get("seance_num"),
        "composite": page1.get("composite"),
        "var_composite": page1.get("var_composite"),
        "brvm30": page1.get("brvm30"),
        "prestige": page1.get("prestige"),
        "nb_actions": nb_actions,
        "actions": actions,
    }


def collect_date(target_date: date, force: bool = False) -> dict | None:
    """Pipeline complet : télécharge + traite un bulletin pour une date."""
    if target_date.weekday() >= 5:
        log.info(f"{target_date} est un week-end, ignoré")
        return None
    
    pdf_path = download_bulletin(target_date, force)
    if not pdf_path:
        return None
    
    return process_bulletin(pdf_path, target_date)


def collect_range(start: date, end: date, force: bool = False) -> list[dict]:
    """Collecte les bulletins pour une plage de dates."""
    results = []
    current = start
    while current <= end:
        if current.weekday() < 5:
            result = collect_date(current, force)
            if result:
                results.append(result)
        current += timedelta(days=1)
    return results


# ── INTERFACE EN LIGNE DE COMMANDE ────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="BRVM Collector v2")
    parser.add_argument("--date", help="Date unique (YYYY-MM-DD)")
    parser.add_argument("--from", dest="from_date", help="Début de plage (YYYY-MM-DD)")
    parser.add_argument("--to", dest="to_date", help="Fin de plage (YYYY-MM-DD)")
    parser.add_argument("--pdf", help="Traiter un fichier PDF local directement")
    parser.add_argument("--force", action="store_true", help="Re-télécharger même si déjà présent")
    parser.add_argument("--excel", action="store_true", help="Générer le rapport Excel")
    parser.add_argument("--summary", action="store_true", help="Afficher le résumé")
    args = parser.parse_args()
    
    init_db()
    
    # Traitement d'un PDF local
    if args.pdf:
        pdf_path = Path(args.pdf)
        if not pdf_path.exists():
            log.error(f"Fichier introuvable : {pdf_path}")
            return
        target_date = date.today()
        if args.date:
            target_date = datetime.strptime(args.date, "%Y-%m-%d").date()
        result = process_bulletin(pdf_path, target_date)
        print(f"\n✓ Traitement terminé : {result['nb_actions']} actions pour {result['date']}")
        print(f"  BRVM Composite : {result['composite']} ({result['var_composite']:+.2f}%)")
        print(f"  BRVM 30        : {result['brvm30']}")
        print(f"  BRVM Prestige  : {result['prestige']}")
        if result['actions']:
            print(f"\n  Top Hausses :")
            top = sorted([a for a in result['actions'] if a.get('variation_jour')], 
                        key=lambda x: x['variation_jour'], reverse=True)[:3]
            for a in top:
                print(f"    {a['symbole']:8} {a.get('cours_cloture', 'N/A'):>10} FCFA  {a['variation_jour']:+.2f}%")
        if args.excel:
            excel_path = generate_excel(datetime.strptime(result['date'], "%Y-%m-%d").date())
            print(f"\n  Rapport Excel : {excel_path}")
        return
    
    # Plage de dates
    if args.from_date:
        start = datetime.strptime(args.from_date, "%Y-%m-%d").date()
        end = datetime.strptime(args.to_date, "%Y-%m-%d").date() if args.to_date else date.today()
        results = collect_range(start, end, args.force)
        print(f"\n✓ {len(results)} bulletins traités")
        return
    
    # Date unique ou aujourd'hui
    target = datetime.strptime(args.date, "%Y-%m-%d").date() if args.date else date.today()
    result = collect_date(target, args.force)
    
    if result:
        print(f"\n✓ {result['nb_actions']} actions collectées pour {result['date']}")
        if args.excel:
            excel_path = generate_excel(target)
            print(f"  Rapport : {excel_path}")
    
    if args.summary:
        conn = sqlite3.connect(DB_PATH)
        nb_seances = conn.execute("SELECT COUNT(*) FROM seances").fetchone()[0]
        nb_cours = conn.execute("SELECT COUNT(*) FROM cours").fetchone()[0]
        conn.close()
        print(f"\n📊 Base de données : {nb_seances} séances, {nb_cours} entrées")


if __name__ == "__main__":
    main()
